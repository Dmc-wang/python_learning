# src/03_formatting.py
"""
学习目标：
1. openpyxl的Workbook/Worksheet对象
2. 单元格样式设置
3. 列宽行高调整
4. 条件格式（进阶）
"""


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import pandas as pd


def create_formatted_report(df: pd.DataFrame, output_path: str):
    """
    创建带格式的Excel报表（学习：对象属性操作）

    知识点：
    - 从pandas到openpyxl的转换
    - 对象实例化：Font(), PatternFill()
    - 颜色表示：RGB和主题色
    - 枚举类型：border_style的选择

    练习：添加边框样式
    """

    # 将DataFrame写入Excel（基础版本）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='数据报表', index=False)

        # 获取workbook和worksheet对象（学习：对象层级关系）
        workbook = writer.book
        worksheet = writer.sheets['数据报表']

        # 设置标题行样式（学习：for循环和枚举）
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 调整列宽（学习：字符串长度计算）
        for idx, col in enumerate(df.columns):
            max_length = max(df[col].astype(str).apply(len).max(), len(col))
            worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

        # 添加条件格式（学习：规则应用）
        # 高亮销售额大于平均值的单元格
        # 练习：修改规则为高亮前3名的值

    print(f"📊 格式化报表已生成: {output_path}")


# 边框设置示例（练习模块）
def add_borders_example():
    """练习：为单元格添加边框"""
    from openpyxl import load_workbook

    wb = Workbook()
    ws = wb.active

    # 合并单元格（学习：合并后只保留左上角数据）
    ws.merge_cells('A1:D1')
    ws['A1'] = "合并标题"

    # 创建边框对象（学习：Side类）
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用边框（练习：循环应用）
    # for row in ws['A1:D5']:
    #     for cell in row:
    #         cell.border = thin_border

    wb.save("data/border_example.xlsx")


# 学习检查点
if __name__ == "__main__":
    import pandas as pd

    data = {'产品': ['A', 'B', 'C'], '销量': [100, 150, 80]}
    df = pd.DataFrame(data)
    create_formatted_report(df, r"D:\Automation_Test_Engineer\python_learning\Dpractice4\data\processed\formatted_report.xlsx")